import streamlit as st
import sqlite3
import pandas as pd
import os

# Configuración estética de la interfaz
st.set_page_config(page_title="Control de Causas - Magaly", layout="wide", page_icon="⚖️")

# Inyección de estilos UI (Tarjetas corporativas azul marino #1B365D)
st.markdown("""
    <style>
    [data-testid="stMetric"] {
        background-color: #F4F7FA;
        padding: 18px;
        border-radius: 12px;
        border-left: 6px solid #1B365D;
        box-shadow: 0px 2px 5px rgba(0,0,0,0.05);
    }
    .stTabs [data-baseweb="tab"] {
        font-size: 13pt;
        font-weight: bold;
        color: #1B365D;
    }
    </style>
""", unsafe_allow_html=True)

st.title("⚖️ Escritorio Jurídico - Control Digital de Causas")
st.markdown("Plataforma interactiva automatizada mediante sondeo local de datos en tiempo real.")

DB_NAME = "C:/Users/Usuario/Documents/DASBOARD_MAGALY/datos.db"
EXCEL_NAME = "C:/Users/Usuario/Documents/DASBOARD_MAGALY/Control_de_Causas_Digital.xlsx"

def conectar_db():
    return sqlite3.connect(DB_NAME)

# Inicialización estructural basada en tus archivos reales
def inicializar_tablas():
    conn = conectar_db()
    cursor = conn.cursor()
    
    # 1. Estructura para Trámites Administrativos
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS administrativos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            cliente TEXT,
            tipo_tramite TEXT,
            recaudos_recibidos TEXT,
            tramites_realizados TEXT,
            estatus TEXT,
            registrado_por TEXT
        )
    ''')
    
    # 2. Estructura para Trámites Tribunalicios
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tribunalicios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            circuito TEXT,
            numero_expediente TEXT,
            tribunal TEXT,
            partes TEXT,
            recurso TEXT,
            actuaciones TEXT,
            observaciones TEXT,
            estatus TEXT,
            registrado_por TEXT
        )
    ''')
    conn.commit()
    conn.close()

inicializar_tablas()

# --- FUNCIÓN DE RESPALDO Y EXPORTACIÓN SEGURA A EXCEL ---
def exportar_a_excel_maestro(df_admin_raw, df_trib_raw):
    df_a = df_admin_raw.copy()
    df_t = df_trib_raw.copy()

    # Intentamos exportar usando openpyxl si está disponible, sino usamos la exportación básica nativa
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        font_title = Font(name="Calibri", size=15, bold=True, color="1B365D")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        thin_border = Border(left=Side(style="thin", color="D3D3D3"), right=Side(style="thin", color="D3D3D3"), top=Side(style="thin", color="D3D3D3"), bottom=Side(style="thin", color="D3D3D3"))
        
        # Hoja Administrativa
        ws1 = wb.active
        ws1.title = "Trámites Administrativos"
        ws1.views.sheetView[0].showGridLines = True
        ws1['A1'] = "CONTROL DE TRÁMITES ADMINISTRATIVOS"
        ws1['A1'].font = font_title
        
        h_admin = ["ID", "Fecha Registro", "Cliente", "Tipo de Trámite", "Recaudos Recibidos", "Trámites Realizados", "Estado", "Registrado Por"]
        for col_idx, text in enumerate(h_admin, 1):
            cell = ws1.cell(row=3, column=col_idx, value=text)
            cell.font = font_header; cell.fill = fill_header; cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, row in enumerate(df_a.values, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=val)
                cell.font = font_data; cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if c_idx in [1, 2, 7] else "left", vertical="center")

        # Hoja Tribunalicia
        ws2 = wb.create_sheet(title="Trámites Tribunalicios")
        ws2.views.sheetView[0].showGridLines = True
        ws2['A1'] = "CONTROL DE EXPEDIENTES Y CAUSAS TRIBUNALICIAS"
        ws2['A1'].font = font_title
        
        h_trib = ["ID", "Fecha Registro", "Circuito", "N° Expediente", "Tribunal", "Partes", "Recurso", "Actuaciones", "Observaciones", "Estado", "Registrado Por"]
        for col_idx, text in enumerate(h_trib, 1):
            cell = ws2.cell(row=3, column=col_idx, value=text)
            cell.font = font_header; cell.fill = fill_header; cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, row in enumerate(df_t.values, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.font = font_data; cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if c_idx in [1, 2, 4, 10] else "left", vertical="center")

        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.row == 1: continue
                    if cell.value: max_len = max(max_len, len(str(cell.value)))
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
        wb.save(EXCEL_NAME)
    except ImportError:
        # CONTINGENCIA: Si no se puede importar openpyxl, guardamos el Excel básico sin estilos usando Pandas
        with pd.ExcelWriter(EXCEL_NAME) as writer:
            df_a.to_excel(writer, sheet_name="Trámites Administrativos", index=False)
            df_t.to_excel(writer, sheet_name="Trámites Tribunalicios", index=False)

# --- PANEL DE BÚSQUEDA LATERAL (UX) ---
st.sidebar.markdown("### 🔍 Panel de Búsqueda Integrado")
buscar_global = st.sidebar.text_input("Buscar por Nombre, Cliente o Expediente:")
estado_seleccionado = st.sidebar.selectbox("Filtrar por Estado actual:", ["Todos", "Activo", "Pendiente", "Sentenciado", "Terminado", "Cerrado"])

# --- RENDERIZADO ASÍNCRONO CADA 5 SEGUNDOS ---
@st.fragment(run_every=5)
def dibujar_interfaz_usuario():
    conn = conectar_db()
    df_admin = pd.read_sql_query("SELECT * FROM administrativos ORDER BY id DESC", conn)
    df_tribunal = pd.read_sql_query("SELECT * FROM tribunalicios ORDER BY id DESC", conn)
    conn.close()

    # Homologación de columnas para evitar fallos de inconsistencia de nombres
    for df in [df_admin, df_tribunal]:
        if not df.empty:
            if 'estatus' in df.columns:
                df.rename(columns={'estatus': 'Estado'}, inplace=True)
            elif 'Estado' not in df.columns:
                df['Estado'] = 'Activo'
                
            if 'fecha_registro' in df.columns:
                df['fecha_registro'] = df['fecha_registro'].astype(str).str.slice(0, 10)

    # Intentar exportar a Excel en segundo plano
    try:
        exportar_a_excel_maestro(df_admin, df_tribunal)
    except Exception:
        pass

    # Aplicación de filtros interactivos
    if buscar_global:
        if not df_admin.empty and 'cliente' in df_admin.columns:
            df_admin = df_admin[df_admin['cliente'].str.contains(buscar_global, case=False, na=False)]
        if not df_tribunal.empty and 'partes' in df_tribunal.columns:
            df_tribunal = df_tribunal[df_tribunal['partes'].str.contains(buscar_global, case=False, na=False) | df_tribunal['numero_expediente'].str.contains(buscar_global, case=False, na=False)]
        
    if estado_seleccionado != "Todos":
        if not df_admin.empty:
            df_admin = df_admin[df_admin['Estado'] == estado_seleccionado]
        if not df_tribunal.empty:
            df_tribunal = df_tribunal[df_tribunal['Estado'] == estado_seleccionado]

    # SECCIÓN DE MÓDULOS DE CONTROL (Tarjetas KPI)
    cant_admin = len(df_admin)
    cant_trib = len(df_tribunal)
    cant_activos = 0
    if not df_admin.empty: cant_activos += len(df_admin[df_admin['Estado'] == 'Activo'])
    if not df_tribunal.empty: cant_activos += len(df_tribunal[df_tribunal['Estado'] == 'Activo'])

    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric(label="🏢 Trámites Administrativos", value=cant_admin)
    with m2:
        st.metric(label="🏛️ Causas en Tribunales", value=cant_trib)
    with m3:
        st.metric(label="🟢 Total Casos Activos", value=cant_activos)

    st.markdown("---")

    # MÓDULO DE PESTAÑAS E INTERFACES TABULARES
    tab_admin, tab_trib = st.tabs(["🏢 Historial Administrativo", "🏛️ Expedientes Tribunalicios"])

    with tab_admin:
        if not df_admin.empty:
            st.dataframe(
                df_admin, 
                use_container_width=True, 
                hide_index=True,
                column_config={
                    "id": "ID", "fecha_registro": "Fecha Registro", "cliente": "Cliente",
                    "tipo_tramite": "Tipo de Trámite", "recaudos_recibidos": "Recaudos Recibidos",
                    "tramites_realizados": "Trámites Realizados", "Estado": "Estado", "registrado_por": "Usuario"
                }
            )
        else:
            st.info("No se registran trámites administrativos bajo los parámetros seleccionados.")

    with tab_trib:
        if not df_tribunal.empty:
            st.dataframe(
                df_tribunal, 
                use_container_width=True, 
                hide_index=True,
                column_config={
                    "id": "ID", "fecha_registro": "Fecha Registro", "circuito": "Circuito",
                    "numero_expediente": "N° Expediente", "tribunal": "Tribunal", "partes": "Partes (Dte vs Ddo)",
                    "recurso": "Recurso", "actuaciones": "Actuaciones", "observaciones": "Observaciones",
                    "Estado": "Estado", "registrado_por": "Usuario"
                }
            )
        else:
            st.info("No se registran causas tribunalicias bajo los parámetros seleccionados.")

    # ACCESO DIRECTO AL ARCHIVO DE RESPALDO EXCEL
    st.markdown("### 📥 Descarga de Archivos Matrices")
    if os.path.exists(EXCEL_NAME):
        with open(EXCEL_NAME, "rb") as file:
            st.download_button(
                label="📥 Descargar Reporte Consolidado en Excel (.xlsx)",
                data=file,
                file_name="Control_de_Causas_Magaly.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

dibujar_interfaz_usuario()